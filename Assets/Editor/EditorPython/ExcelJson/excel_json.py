import sys
import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def resolve_path(path_setting):
    """解析路径（支持路径索引）"""
    if path_setting and "." in path_setting:
        # 这里可以添加路径索引解析逻辑
        # 暂时直接返回路径设置
        return path_setting
    return path_setting

def get_relative_path(from_path, to_path):
    """计算相对路径（兼容方式）"""
    try:
        return os.path.relpath(to_path, from_path)
    except ValueError:
        # 如果不包含关系，返回文件名
        return os.path.basename(to_path)

def excel_to_json(excel_path, json_path):
    """将单个Excel文件转换为JSON"""
    try:
        # 读取Excel文件
        workbook = load_workbook(excel_path, data_only=True)
        sheet = workbook.active  # 使用第一个工作表

        # 读取表头（第1行）
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        columns = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(header_row)]

        # 读取类型行（第2行，可选）
        type_row = None
        try:
            type_rows = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
            if type_rows:
                type_row = [str(cell).lower().strip() if cell is not None else "" for cell in type_rows[0]]
        except:
            pass

        # 读取数据行（从第3行开始，如果有类型行的话）
        data_start_row = 3 if type_row else 2
        json_data = []

        for row_idx in range(data_start_row, sheet.max_row + 1):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            row_dict = {}
            has_data = False

            for col_idx, value in enumerate(row):
                if col_idx >= len(columns):
                    continue

                # 检查是否有数据
                if value is not None and str(value).strip():
                    has_data = True

                # 类型转换
                if type_row and col_idx < len(type_row):
                    type_str = type_row[col_idx]
                    converted_value = convert_value_by_type(str(value) if value is not None else "", type_str)
                else:
                    converted_value = convert_value_auto(value)

                row_dict[columns[col_idx]] = converted_value

            # 只有当行有数据时才添加到结果
            if has_data:
                json_data.append(row_dict)

        workbook.close()
        return json_data

    except Exception as e:
        raise Exception(f"转换Excel文件失败 {excel_path}: {str(e)}")

def convert_value_by_type(value_str, type_str):
    """根据类型字符串转换值"""
    if not value_str or value_str.strip() == '' or value_str.lower() == 'nan':
        return None

    value_str = value_str.strip()

    try:
        if type_str == 'int':
            return int(float(value_str))
        elif type_str == 'float':
            return float(value_str)
        elif type_str == 'bool':
            return value_str.lower() in ('true', '1', 'yes', 'on')
        elif type_str == 'int[]':
            return [int(float(x.strip())) for x in value_str.split(',') if x.strip()]
        elif type_str == 'float[]':
            return [float(x.strip()) for x in value_str.split(',') if x.strip()]
        elif type_str == 'string[]':
            return [x.strip() for x in value_str.split(',') if x.strip()]
        else:
            return value_str
    except:
        return value_str

def convert_value_auto(value):
    """自动推断值类型"""
    if value is None:
        return None

    value_str = str(value).strip()

    # 检查是否包含逗号（可能是数组）
    if ',' in value_str:
        parts = [x.strip() for x in value_str.split(',') if x.strip()]
        if not parts:
            return []

        # 尝试转换为数字数组
        try:
            # 尝试int数组
            int_array = [int(float(x)) for x in parts]
            return int_array
        except:
            try:
                # 尝试float数组
                float_array = [float(x) for x in parts]
                return float_array
            except:
                # 返回字符串数组
                return parts

    # 尝试转换为数字
    try:
        # 尝试int
        if '.' not in value_str:
            return int(float(value_str))
        else:
            return float(value_str)
    except:
        pass

    # 尝试转换为布尔值
    if value_str.lower() in ('true', 'false', '1', '0', 'yes', 'no', 'on', 'off'):
        return value_str.lower() in ('true', '1', 'yes', 'on')

    # 默认返回字符串
    return value_str

def json_to_excel(json_data, excel_path):
    """将JSON数据转换为Excel文件"""
    try:
        if not json_data or len(json_data) == 0:
            raise Exception("JSON数据为空")

        # 创建工作簿
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'

        # 获取字段名
        if len(json_data) > 0:
            first_obj = json_data[0]
            keys = list(first_obj.keys())

            # 第一行：字段名
            for col_idx, key in enumerate(keys):
                sheet.cell(row=1, column=col_idx+1, value=key)

            # 第二行：类型推断
            for col_idx, key in enumerate(keys):
                value = first_obj[key]
                type_str = infer_type_from_json(value)
                sheet.cell(row=2, column=col_idx+1, value=type_str)

            # 从第三行开始写入数据
            for row_idx, obj in enumerate(json_data):
                for col_idx, key in enumerate(keys):
                    value = obj.get(key, "")
                    if isinstance(value, list):
                        value = ",".join(str(item) for item in value)
                    sheet.cell(row=row_idx+3, column=col_idx+1, value=value)

        # 保存文件
        workbook.save(excel_path)
        workbook.close()
        return True

    except Exception as e:
        raise Exception(f"转换JSON到Excel失败 {excel_path}: {str(e)}")

def infer_type_from_json(value):
    """从JSON值推断类型"""
    if value is None:
        return "string"

    if isinstance(value, bool):
        return "bool"
    elif isinstance(value, int):
        return "int"
    elif isinstance(value, float):
        return "float"
    elif isinstance(value, list):
        if not value:
            return "string[]"
        first_item = value[0]
        if isinstance(first_item, int):
            return "int[]"
        elif isinstance(first_item, float):
            return "float[]"
        else:
            return "string[]"
    else:
        return "string"

def convert_all_excel_to_json(excel_folder, json_folder):
    """批量将Excel文件转换为JSON"""
    excel_path = resolve_path(excel_folder)
    json_path = resolve_path(json_folder)

    if not os.path.exists(excel_path):
        raise Exception(f"Excel文件夹不存在: {excel_path}")

    total_converted = 0
    total_failed = 0

    # 递归查找所有xlsx文件
    excel_files = []
    for root, dirs, files in os.walk(excel_path):
        for file in files:
            if file.endswith('.xlsx'):
                excel_files.append(os.path.join(root, file))

    for excel_file in excel_files:
        try:
            # 计算相对路径
            relative_path = get_relative_path(excel_path, excel_file)
            relative_dir = os.path.dirname(relative_path)
            file_name_without_ext = os.path.splitext(os.path.basename(excel_file))[0]

            # 构建JSON文件路径
            json_dir = os.path.join(json_path, relative_dir) if relative_dir else json_path
            os.makedirs(json_dir, exist_ok=True)
            json_file = os.path.join(json_dir, f"{file_name_without_ext}.json")

            # 转换Excel到JSON
            json_data = excel_to_json(excel_file, json_file)
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)

            total_converted += 1
            json_relative_path = get_relative_path(json_path, json_file)
            print(f"✅ 转换完成: {relative_path} → {json_relative_path}")

        except Exception as e:
            total_failed += 1
            print(f"❌ 转换失败: {excel_file}\n{str(e)}")

    print(f"✅ Excel → JSON 批量转换完成！成功: {total_converted} 个，失败: {total_failed} 个")

def convert_all_json_to_excel(json_folder, excel_folder):
    """批量将JSON文件转换为Excel"""
    json_path = resolve_path(json_folder)
    excel_path = resolve_path(excel_folder)

    if not os.path.exists(json_path):
        raise Exception(f"JSON文件夹不存在: {json_path}")

    total_converted = 0
    total_failed = 0

    # 递归查找所有json文件
    json_files = []
    for root, dirs, files in os.walk(json_path):
        for file in files:
            if file.endswith('.json'):
                json_files.append(os.path.join(root, file))

    for json_file in json_files:
        try:
            # 计算相对路径
            relative_path = get_relative_path(json_path, json_file)
            relative_dir = os.path.dirname(relative_path)
            file_name_without_ext = os.path.splitext(os.path.basename(json_file))[0]

            # 构建Excel文件路径
            excel_dir = os.path.join(excel_path, relative_dir) if relative_dir else excel_path
            os.makedirs(excel_dir, exist_ok=True)
            excel_file = os.path.join(excel_dir, f"{file_name_without_ext}.xlsx")

            # 读取JSON内容
            with open(json_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)

            # 转换JSON到Excel
            json_to_excel(json_data, excel_file)

            total_converted += 1
            excel_relative_path = get_relative_path(excel_path, excel_file)
            print(f"✅ 转换完成: {relative_path} → {excel_relative_path}")

        except Exception as e:
            total_failed += 1
            print(f"❌ 转换失败: {json_file}\n{str(e)}")

    print(f"✅ JSON → Excel 批量转换完成！成功: {total_converted} 个，失败: {total_failed} 个")

def main():
    if len(sys.argv) < 4:
        print("用法: python excel_json.py <excel_folder> <json_folder> <conversion_mode>")
        print("conversion_mode: 'Excel → JSON', 'JSON → Excel', 或 '双向转换'")
        return

    excel_folder = sys.argv[1]
    json_folder = sys.argv[2]
    conversion_mode = sys.argv[3]

    print("开始ExcelJson转换...")
    print(f"Excel文件夹: {excel_folder}")
    print(f"JSON文件夹: {json_folder}")
    print(f"转换模式: {conversion_mode}")
    print()

    try:
        if conversion_mode == "Excel → JSON":
            print("=== Excel → JSON 转换 ===")
            convert_all_excel_to_json(excel_folder, json_folder)
        elif conversion_mode == "JSON → Excel":
            print("=== JSON → Excel 转换 ===")
            convert_all_json_to_excel(json_folder, excel_folder)
        elif conversion_mode == "双向转换":
            # Excel → JSON
            print("=== Excel → JSON 转换 ===")
            convert_all_excel_to_json(excel_folder, json_folder)

            print()

            # JSON → Excel
            print("=== JSON → Excel 转换 ===")
            convert_all_json_to_excel(json_folder, excel_folder)
        else:
            raise Exception(f"未知的转换模式: {conversion_mode}")

        print()
        print("✅ 转换完成！")

    except Exception as e:
        print(f"❌ 执行失败: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
