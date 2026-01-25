import sys
import os
import re
import json
import warnings
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# 过滤openpyxl的样式警告
warnings.filterwarnings("ignore", message="Workbook contains no default style", category=UserWarning)

# 模拟枚举系统
class LangTypeEnum:
    Chinese = "Chinese"
    English = "English"
    Malayu = "Malayu"

    @staticmethod
    def get_all():
        return [LangTypeEnum.Chinese, LangTypeEnum.English, LangTypeEnum.Malayu]

class LangModuleEnum:
    UI = "UI"
    ItemInfo = "ItemInfo"
    Default = "Default"

    @staticmethod
    def get_all():
        return [LangModuleEnum.UI, LangModuleEnum.ItemInfo, LangModuleEnum.Default]

def resolve_path(path_setting):
    """解析路径（支持路径索引）"""
    if path_setting and "." in path_setting:
        # 这里可以添加路径索引解析逻辑
        # 暂时直接返回路径设置
        return path_setting
    return path_setting

def get_relative_path(from_path, to_path):
    """计算相对路径（兼容方式）"""
    try:
        return os.path.relpath(to_path, from_path)
    except ValueError:
        # 如果不包含关系，返回文件名
        return os.path.basename(to_path)

def scan_project_for_module(assets_path, module_name):
    """扫描项目中指定模块的所有GetLangContent调用"""
    lang_entries = {}

    # 扫描所有C#文件
    cs_files = []
    for root, dirs, files in os.walk(assets_path):
        for file in files:
            if file.endswith('.cs'):
                cs_files.append(os.path.join(root, file))

    # 正则表达式：匹配固定字符串的GetLangContent调用
    regex = re.compile(
        r'GetLangContent\s*\(\s*(?:(?:[\w\.]+\.)?LangModuleEnum\.(\w+)\s*,\s*)?""([^""]+)""\s*\)',
        re.IGNORECASE
    )

    # 正则表达式：匹配注释中的语言键定义
    comment_regex = re.compile(
        r'//\s*GetLangContent\s*:\s*(\w+)\s*,\s*\{([^}]+)\}',
        re.IGNORECASE
    )

    for file_path in cs_files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # 扫描固定字符串的GetLangContent调用
            matches = regex.findall(content)
            for match in matches:
                scanned_module_name = match[0] if match[0] else "Default"
                key = match[1]

                if scanned_module_name.lower() == module_name.lower():
                    if key not in lang_entries:
                        lang_entries[key] = ""

            # 扫描注释中的语言键定义
            comment_matches = comment_regex.findall(content)
            for comment_match in comment_matches:
                scanned_module_name = comment_match[0]
                keys_string = comment_match[1]

                if scanned_module_name.lower() == module_name.lower():
                    # 解析键列表，支持 "str1","str2" 格式
                    keys = [k.strip().strip('"\'') for k in keys_string.split(',') if k.strip()]
                    for key in keys:
                        if key and key not in lang_entries:
                            lang_entries[key] = ""

        except Exception as e:
            print(f"扫描文件失败 {file_path}: {str(e)}")

    return lang_entries

def read_from_excel(excel_path):
    """从Excel读取语言配置"""
    result = {}

    if not os.path.exists(excel_path):
        print(f"语言 Excel 文件不存在: {excel_path}")
        return result

    try:
        # 读取Excel文件，使用只读模式避免样式警告
        workbook = load_workbook(excel_path, data_only=True, read_only=True)
        sheet = workbook.active

        if sheet is None:
            print(f"Excel 文件没有工作表: {excel_path}")
            return result

        # 跳过表头行（第1行）和类型行（第2行），从第3行开始读取数据
        for row_idx in range(3, sheet.max_row + 1):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # 读取 key（第2列，索引1）
            if len(row) < 2 or row[1] is None:
                continue

            key = str(row[1]).strip()
            if not key:
                continue

            # 读取 content（第3列，索引2）
            content = str(row[2]).strip() if len(row) >= 3 and row[2] is not None else ""

            result[key] = content

        workbook.close()
        print(f"从 Excel 读取了 {len(result)} 条语言配置: {excel_path}")

    except Exception as e:
        print(f"读取 Excel 文件失败: {excel_path}\n{str(e)}")

    return result

def update_excel(excel_path, scanned_keys):
    """更新Excel文件（只新建新项，删除没有了的项，不改变现有项的赋值）"""
    try:
        if os.path.exists(excel_path):
            # 加载现有文件时不检查样式
            workbook = load_workbook(excel_path, data_only=False)
            sheet = workbook.active
            if sheet is None:
                sheet = workbook.create_sheet("Sheet1")
        else:
            # 创建新文件
            workbook = Workbook()
            sheet = workbook.active

        # 确保表头和类型行存在
        header_row = sheet.cell(row=1, column=1)
        if not header_row.value:
            sheet.cell(row=1, column=1, value="id")
            sheet.cell(row=1, column=2, value="key")
            sheet.cell(row=1, column=3, value="content")

        type_row = sheet.cell(row=2, column=1)
        if not type_row.value:
            sheet.cell(row=2, column=1, value="int")
            sheet.cell(row=2, column=2, value="string")
            sheet.cell(row=2, column=3, value="string")

        # 读取现有数据（从第3行开始），收集需要保留的数据
        rows_to_keep = {}
        for row_idx in range(3, sheet.max_row + 1):
            key_cell = sheet.cell(row=row_idx, column=2)
            if key_cell.value is None:
                continue

            key = str(key_cell.value).strip()
            if not key:
                continue

            # 如果这个键在扫描结果中，保留现有值
            if key in scanned_keys:
                content_cell = sheet.cell(row=row_idx, column=3)
                existing_value = str(content_cell.value).strip() if content_cell.value is not None else ""
                rows_to_keep[key] = existing_value

        # 添加新项（在扫描结果中但不在 Excel 中）
        for key in scanned_keys:
            if key not in rows_to_keep:
                rows_to_keep[key] = ""  # 新项默认空值

        # 删除所有数据行（从第3行开始）
        for row_idx in range(sheet.max_row, 2, -1):
            sheet.delete_rows(row_idx)

        # 重新写入数据行
        id_counter = 1
        sorted_keys = sorted(rows_to_keep.keys())
        for key in sorted_keys:
            row_idx = id_counter + 2  # 第3行开始
            sheet.cell(row=row_idx, column=1, value=id_counter)
            sheet.cell(row=row_idx, column=2, value=key)
            sheet.cell(row=row_idx, column=3, value=rows_to_keep[key])
            id_counter += 1

        # 保存文件
        workbook.save(excel_path)
        workbook.close()

        print(f"更新 Excel 完成: {excel_path} (共 {len(rows_to_keep)} 条)")

    except Exception as e:
        print(f"更新 Excel 文件失败: {excel_path}\n{str(e)}")

def scan_and_update_all_excels(excel_folder, assets_path):
    """扫描并更新所有Excel文件"""
    excel_path = resolve_path(excel_folder)
    if not os.path.exists(excel_path):
        os.makedirs(excel_path, exist_ok=True)

    # 获取所有语言和模块
    all_languages = LangTypeEnum.get_all()
    all_modules = LangModuleEnum.get_all()

    if not all_languages:
        print("❌ 无法获取语言列表")
        return

    if not all_modules:
        print("❌ 无法获取模块列表")
        return

    total_updated = 0

    # 遍历所有语言和模块组合
    for lang_name in all_languages:
        for module_name in all_modules:
            # 扫描项目获取该模块的所有键
            scanned_keys = scan_project_for_module(assets_path, module_name)

            if not scanned_keys:
                continue

            # 构建 Excel 文件路径
            excel_file_name = f"{lang_name}_{module_name}_Lang.xlsx"
            excel_file_path = os.path.join(excel_path, excel_file_name)

            # 更新 Excel
            update_excel(excel_file_path, scanned_keys)
            total_updated += 1

    print(f"✅ 扫描并更新完成！共更新 {total_updated} 个 Excel 文件")

def convert_all_excels_to_json(excel_folder, json_folder):
    """转换所有语言Excel文件生成JSON"""
    excel_path = resolve_path(excel_folder)
    json_path = resolve_path(json_folder)

    if not os.path.exists(excel_path):
        print(f"Excel 文件夹不存在: {excel_path}")
        return

    if not os.path.exists(json_path):
        os.makedirs(json_path, exist_ok=True)

    # 获取所有语言和模块
    all_languages = LangTypeEnum.get_all()
    all_modules = LangModuleEnum.get_all()

    if not all_languages:
        print("❌ 无法获取语言列表")
        return

    if not all_modules:
        print("❌ 无法获取模块列表")
        return

    total_converted = 0

    # 遍历所有语言和模块组合
    for lang_name in all_languages:
        for module_name in all_modules:
            # 构建 Excel 文件路径
            excel_file_name = f"{lang_name}_{module_name}_Lang.xlsx"
            excel_file_path = os.path.join(excel_path, excel_file_name)

            if not os.path.exists(excel_file_path):
                continue

            # 从 Excel 读取数据
            lang_data = read_from_excel(excel_file_path)

            if not lang_data:
                continue

            # 生成 JSON 文件
            json_file_name = f"{lang_name}_{module_name}_Lang.json"
            json_file_path = os.path.join(json_path, json_file_name)

            with open(json_file_path, 'w', encoding='utf-8') as f:
                json.dump(lang_data, f, ensure_ascii=False, indent=2)

            total_converted += 1
            print(f"✅ 转换完成: {json_file_name} ({len(lang_data)} 条)")

    print(f"✅ 所有语言配置文件转换完成！共转换 {total_converted} 个 JSON 文件")

def main():
    if len(sys.argv) < 4:
        print("用法: python language_configer.py <excel_folder> <json_output_folder> <operation_mode>")
        print("operation_mode: '扫描并更新 Excel', '转换所有语言配置文件生成 JSON'")
        return

    excel_folder = sys.argv[1]
    json_output_folder = sys.argv[2]
    operation_mode = sys.argv[3]

    print("=== Python 语言配置器 ===")
    print(f"Excel文件夹: {excel_folder}")
    print(f"JSON输出文件夹: {json_output_folder}")
    print(f"操作模式: {operation_mode}")
    print()

    try:
        # 获取Assets路径（从脚本位置推断）
        script_dir = os.path.dirname(os.path.abspath(__file__))
        editor_python_path = os.path.dirname(script_dir)
        assets_path = os.path.dirname(editor_python_path)

        if operation_mode == "扫描并更新 Excel":
            print("=== 扫描并更新 Excel ===")
            scan_and_update_all_excels(excel_folder, assets_path)
        elif operation_mode == "转换所有语言配置文件生成 JSON":
            print("=== 转换所有语言配置文件生成 JSON ===")
            convert_all_excels_to_json(excel_folder, json_output_folder)
        else:
            raise Exception(f"未知的操作模式: {operation_mode}")

        print()
        print("✅ 操作完成！")

    except Exception as e:
        print(f"❌ 执行失败: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
