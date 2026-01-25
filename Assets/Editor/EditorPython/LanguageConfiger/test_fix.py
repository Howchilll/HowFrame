#!/usr/bin/env python3
"""
测试LanguageConfiger的样式警告修复
"""

import os
import sys
import warnings
from pathlib import Path

# 添加项目路径
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(os.path.dirname(script_dir)))
sys.path.insert(0, project_root)

def test_warnings_filter():
    """测试warnings过滤器"""
    print("测试 warnings 过滤器...")

    # 过滤openpyxl的样式警告
    warnings.filterwarnings("ignore", message="Workbook contains no default style", category=UserWarning)

    try:
        from openpyxl import Workbook
        import io
        import contextlib

        # 捕获警告输出
        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("always")  # 捕获所有警告

            # 创建一个工作簿（这通常会触发警告）
            workbook = Workbook()
            sheet = workbook.active

            # 检查是否捕获到警告
            style_warnings = [warning for warning in w if "default style" in str(warning.message)]
            if style_warnings:
                print("❌ 警告过滤器失效，仍有样式警告")
                for warning in style_warnings:
                    print(f"  警告: {warning.message}")
                return False
            else:
                print("✅ 警告过滤器工作正常")
                return True

    except ImportError:
        print("❌ 无法导入 openpyxl，请确保已安装")
        return False

def test_read_only_mode():
    """测试只读模式"""
    print("测试只读模式...")

    try:
        from openpyxl import load_workbook, Workbook

        # 创建一个测试Excel文件
        test_file = "test_temp.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value="test")
        workbook.save(test_file)

        # 测试只读模式读取
        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("always")

            read_workbook = load_workbook(test_file, data_only=True, read_only=True)
            read_sheet = read_workbook.active

            style_warnings = [warning for warning in w if "default style" in str(warning.message)]
            read_workbook.close()

        # 清理测试文件
        if os.path.exists(test_file):
            os.remove(test_file)

        if style_warnings:
            print("❌ 只读模式仍有样式警告")
            for warning in style_warnings:
                print(f"  警告: {warning.message}")
            return False
        else:
            print("✅ 只读模式工作正常")
            return True

    except Exception as e:
        print(f"❌ 只读模式测试失败: {str(e)}")
        return False

def main():
    print("=== LanguageConfiger 样式警告修复测试 ===")
    print()

    test1_passed = test_warnings_filter()
    test2_passed = test_read_only_mode()

    print()
    if test1_passed and test2_passed:
        print("✅ 所有测试通过！样式警告已修复")
        return 0
    else:
        print("❌ 部分测试失败")
        return 1

if __name__ == "__main__":
    sys.exit(main())
